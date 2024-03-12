import bpy
import pandas as pd
from openpyxl import load_workbook



# Specify the path to your Excel file
excel_file_path = "D://projects/NPC/20221117 3d model/automation phase/new naming templates/namin.xlsx"
sheet_name = "meshes"  # Replace with the actual sheet name in your Excel file
filter_column_name = "category"  # Replace with the actual column name in your Excel file
target_column_name = "objects"  # Replace with the actual column name containing object names

# Load the Excel workbook with openpyxl
wb = load_workbook(excel_file_path, read_only=True, data_only=True)
ws = wb[sheet_name]

# Get all data from the Excel sheet
header = None
all_data = []

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    if header is None:
        header = [cell.value for cell in row]
    else:
        row_data = {header[i]: cell.value for i, cell in enumerate(row)}
        all_data.append(row_data)

# Convert the data to a DataFrame
df = pd.DataFrame(all_data)

# Specify the filter values
filter_values = ["foundation"]

# Filter the DataFrame based on the specified filter values
filtered_data = df[df[filter_column_name].isin(filter_values)]
object_names_to_show = filtered_data[target_column_name].tolist()

# Get the active scene
scene = bpy.context.scene

# Reset visibility of all objects
for obj in scene.objects:
    obj.hide_viewport = False

# Iterate through all objects in the scene
for obj in scene.objects:
    # Check if the object name is in the list from the Excel file
    if obj.name in object_names_to_show:
        # Show the object
        obj.hide_viewport = False
    else:
        # Hide the object
        obj.hide_viewport = True

# Close the Excel workbook to release resources
wb.close()

print("Objects visibility updated based on the filtered Excel column.")
