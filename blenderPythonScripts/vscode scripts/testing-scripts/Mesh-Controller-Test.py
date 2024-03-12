import bpy
import pandas as pd
from openpyxl import load_workbook

def load_excel_data(excel_file_path, sheet_name):
    wb = load_workbook(excel_file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    header = None
    all_data = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if header is None:
            header = [cell.value for cell in row]
        else:
            row_data = {header[i]: cell.value for i, cell in enumerate(row)}
            all_data.append(row_data)

    wb.close()
    return pd.DataFrame(all_data)

def filter_objects_by_column(data_frame, filter_column_name, filter_values, target_column_name):
    filtered_data = data_frame[data_frame[filter_column_name].isin(filter_values)]
    return filtered_data[target_column_name].tolist()

def update_object_visibility(object_names_to_show):
    scene = bpy.context.scene

    for obj in scene.objects:
        obj.hide_set(obj.name not in object_names_to_show)

def main():
    # Specify the path to your Excel file
    excel_file_path = "D://projects/NPC/20221117 3d model/automation phase/new naming templates/namin.xlsx"
    sheet_name = "meshes"  # Replace with the actual sheet name in your Excel file
    filter_column_name = "category"  # Replace with the actual column name in your Excel file
    target_column_name = "objects"  # Replace with the actual column name containing object names

    # Load Excel data
    df = load_excel_data(excel_file_path, sheet_name)

    # Specify the filter values
    filter_values = ["duplicated"]

    # Filter the DataFrame based on the specified filter values
    object_names_to_show = filter_objects_by_column(df, filter_column_name, filter_values, target_column_name)

    # Update object visibility in Blender
    update_object_visibility(object_names_to_show)

    print("Objects visibility updated based on the filtered Excel column.")

# Run the main function
main()
