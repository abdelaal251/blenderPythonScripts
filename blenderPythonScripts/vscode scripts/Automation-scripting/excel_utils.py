import pandas as pd
from openpyxl import load_workbook


def load_excel_data(excel_file_path, sheet_name):
    wb = load_workbook(excel_file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    header = None
    all_data = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if header is None:
            header = [cell.value for cell in row]
        else:
            row_data = {header[i]: cell.value for i, cell in enumerate(row)}
            all_data.append(row_data)

    wb.close()
    return pd.DataFrame(all_data)


def filter_objects_by_column(df, filter_column_name, filter_values, target_column_name):
    filtered_data = df[df[filter_column_name].isin(filter_values)]
    return filtered_data[target_column_name].tolist()
